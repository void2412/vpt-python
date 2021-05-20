import sys
from dataclasses import dataclass
import openpyxl
from PyQt5.QtWidgets import *
from gui import mainMenu
from mapData import *
import startvpt
from openpyxl import Workbook
from PyQt5.QtCore import *
import subprocess


@dataclass
class account:
    title: str
    link: str
    group: str


class mainMenuUI(QMainWindow, mainMenu.Ui_mainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.dailyWin = []
        self.counter = 0
        self.resizeTable()
        self.saveBtn.clicked.connect(self.SaveAccount)
        self.startBtn.clicked.connect(self.startBtnClicked)
        self.addBtn.clicked.connect(self.addRow)
        self.delBtn.clicked.connect(self.removeRow)
        self.accounts = []
        self.groups = []
        self.saveacc = []
        self.listWidget.itemClicked.connect(self.listItemClicked)
        self.tableWidget.setSortingEnabled(False)
        self.populateData()
        self.autoKsBtn.clicked.connect(self.moAutoKs)
        self.tabWidget.setCurrentIndex(0)

    def moAutoKs(self):
        subprocess.Popen('autoKs.exe')

    def populateData(self):
        self.accounts.clear()
        self.groups.clear()
        self.groups = ['All']
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(0)
        self.listWidget.clear()
        workbookData = openpyxl.load_workbook(filename='./account/link_account.xlsx')
        sheet = workbookData.active
        for row in sheet.iter_rows(min_row=2, values_only=True):

            if (row[title] != None):
                t = row[title].strip()
            else:
                t = ''
            if (row[link] != None):
                l = row[link].strip()
            else:
                l = ''
            if (row[group] != None):
                g = row[group].strip()
            else:
                g = ''

            acc = account(title=t.strip(),
                          link=l.strip(),
                          group=g.strip())
            if acc.title!='' or acc.link!='' or acc.group!='':
                self.accounts.append(acc)
            self.groups.append(acc.group)


        self.groups = list(dict.fromkeys(self.groups))
        for item in self.groups:
            if (item != ''):
                self.listWidget.addItem(item)
        self.listWidget.addItem('No Group')
        self.listWidget.setCurrentRow(0)

        self.showData('All')

    def listItemClicked(self):
        x = self.listWidget.selectedItems()
        groupName = x[0].text()
        self.showData(groupName)

    def showData(self, groupName):
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(0)
        index = 0
        for acc in self.accounts:
            if groupName == 'All':
                self.addAccountToTable(index, acc.title, acc.link, acc.group)
                index = index + 1

            elif groupName == 'No Group':
                if acc.group == '':
                    self.addAccountToTable(index, acc.title, acc.link, acc.group)
                    index = index + 1
            else:
                if acc.group == groupName:
                    self.addAccountToTable(index, acc.title, acc.link, acc.group)
                    index = index + 1

        pass

    def resizeTable(self):
        x = self.tableWidget.horizontalHeader()
        x.setSectionResizeMode(1, QHeaderView.Stretch)

    def addAccountToTable(self, rowPos, title, link, group):
        rowCount = self.tableWidget.rowCount()
        self.tableWidget.insertRow(rowCount)
        self.tableWidget.setItem(rowPos, 0, QTableWidgetItem(title))
        self.tableWidget.setItem(rowPos, 1, QTableWidgetItem(link))
        self.tableWidget.setItem(rowPos, 2, QTableWidgetItem(group))

    def SaveAccount(self):
        self.saveacc.clear()
        for row in range(self.tableWidget.rowCount()):
            t = self.tableWidget.item(row, 0).text()
            l = self.tableWidget.item(row, 1).text()
            g = self.tableWidget.item(row, 2).text()
            acc = account(title=t, link=l, group=g)
            self.saveacc.append(acc)

        for saveacc in self.saveacc:
            exist = False
            for acc in self.accounts:
                if acc.link == saveacc.link:
                    exist = True
                    acc.title = saveacc.title
                    acc.group = saveacc.group
                    break
                if acc.title == saveacc.title:
                    exist = True
                    acc.link = saveacc.link
                    acc.group = saveacc.group
            if exist == False:
                self.accounts.append(saveacc)


        toSave = Workbook()
        sheet = toSave.active
        index = 2
        sheet['A1'] = 'title'
        sheet['B1'] = 'link'
        sheet['C1'] = 'group'
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 170
        sheet.column_dimensions['C'].width = 20
        for data in self.accounts:
            if (data.title != '' or data.link != '' or data.group != ''):
                sheet.cell(row=index, column=1).value = data.title.strip()
                sheet.cell(row=index, column=2).value = data.link.strip()
                sheet.cell(row=index, column=3).value = data.group.strip()
            index = index + 1
        sheet.auto_filter.ref = sheet.dimensions
        toSave.save('./account/link_account.xlsx')
        self.populateData()
        pass

    def startBtnClicked(self):
        selected = self.tableWidget.selectionModel().selectedRows()

        for row in selected:
            title = self.tableWidget.item(row.row(), 0).text()
            link = self.tableWidget.item(row.row(), 1).text()
            startvpt.startGame(title, link)
        pass

    def addRow(self):
        rowCount = self.tableWidget.rowCount()
        self.addAccountToTable(rowCount, '', '', '')
        self.tableWidget.selectRow(len(self.accounts))
        pass

    def removeRow(self):

        indexes = [QPersistentModelIndex(index) for index in self.tableWidget.selectionModel().selectedRows()]
        for index in indexes:
            t = self.tableWidget.item(index.row(),0).text()
            l = self.tableWidget.item(index.row(),1).text()
            g = self.tableWidget.item(index.row(),2).text()
            acc = account(t,l,g)
            for i in range(len(self.accounts)):
                if self.accounts[i] == acc:
                    self.accounts.pop(i)
                    break
            self.tableWidget.removeRow(index.row())


if __name__ == '__main__':
    app = QApplication(sys.argv)
    mainMenu = mainMenuUI()
    mainMenu.show()
    sys.exit(app.exec_())
